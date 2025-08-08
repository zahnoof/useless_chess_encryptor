# In your excel_manager.py file

import openpyxl

def save_game_position(board):
    """Saves the current board state to an Excel file."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Save the board state to the Excel sheet
    for r, row in enumerate(board):
        for c, piece in enumerate(row):
            sheet.cell(row=r + 1, column=c + 1, value=piece)

    workbook.save("game_position.xlsx")
    print("Game position saved to game_position.xlsx")

def load_game_position(filename):
    """Loads a board state from an Excel file."""
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        board = []
        for r in range(1, 9):
            row = []
            for c in range(1, 9):
                row.append(sheet.cell(row=r, column=c).value)
            board.append(row)
        return board
    except FileNotFoundError:
        print(f"Error: The file '{filename}' was not found.")
        return None